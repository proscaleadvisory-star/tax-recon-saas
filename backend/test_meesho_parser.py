import io
from unittest.mock import MagicMock
import pytest
import pandas as pd
import openpyxl
from sqlalchemy.orm import Session

from meesho_contracts import MEESHO_CONTRACTS, ReportContractSchema
from extractor import MeeshoExtractionEngine, parse_date, clean_numeric

# Mark synthetic test data explicitly as requested
SYNTHETIC_LABEL = "SYNTHETIC_TEST_FIXTURE_NOT_REAL_MARKETPLACE_EXPORT"

def create_synthetic_excel(sheets_data: dict) -> bytes:
    """Helper to create an in-memory workbook with multiple sheets."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

def test_workbook_classifier():
    """Verify that file names and sheets are correctly classified under Meesho contracts."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")

    # 1. Payout report classification
    contract = engine.classify_file(
        "SP_ORDER_ADS_REFERRAL_PAYMENT_FILE_PREVIOUS_PAYMENT_20260601.xlsx",
        ["Order Payments", "Disclaimer", "Ads Cost", "Referral Payments", "Compensation and Recovery"]
    )
    assert contract is not None
    assert contract.id == "MEESHO_PAYMENT_PREVIOUS_PAYMENT_CURRENT"
    assert contract.report_family == "payout"

    # 2. Sales report classification
    contract = engine.classify_file("Meesho outward sales register June.xlsx", ["sales report format"])
    assert contract is not None
    assert contract.id == "MEESHO_GST_TCS_SALES_CURRENT"
    
    # 3. Credit note / returns classification
    contract = engine.classify_file("Meesho return credit note report.xlsx", ["sales report format"])
    assert contract is not None
    assert contract.id == "MEESHO_GST_TCS_SALES_RETURN_CURRENT"

    # 4. Unknown Meesho file
    contract = engine.classify_file("meesho_some_other_sheet.xlsx", ["Sheet1"])
    assert contract is None

def test_unknown_meesho_format_rejection():
    """Verify that process_and_save rejects unrecognized formats with Meesho in their name."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")
    
    # Mock create_failed_upload
    engine.create_failed_upload = MagicMock()
    
    sheets_data = {
        "Sheet1": [
            ["Column A", "Column B"],
            ["Val A", "Val B"]
        ]
    }
    content = create_synthetic_excel(sheets_data)
    
    with pytest.raises(ValueError) as exc_info:
        engine.process_and_save(content, "Unrecognized_Meesho_Format_2026.xlsx", "2026-06")
        
    assert "failed classification" in str(exc_info.value)
    engine.create_failed_upload.assert_called_once()



def test_gst_horizontal_block_slicing():
    """Verify that Outward supplies and returns blocks are sliced correctly from one horizontal sheet."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")

    # Side-by-side tabular dataframe
    # Section 0 (Sales) vs Section 1 (Credit Notes / CN.1)
    df_raw = pd.DataFrame({
        "sub_order_num": ["123", "456"],
        "order_date": ["2026-06-01", "2026-06-02"],
        "total_taxable_sale_value": ["100.0", "200.0"],
        "total_invoice_value": ["118.0", "236.0"],
        "sub_order_num.1": ["CN-123", "CN-456"],
        "order_date.1": ["2026-06-05", "2026-06-06"],
        "total_taxable_sale_value.1": ["-10.0", "-20.0"],
        "total_invoice_value.1": ["-11.8", "-23.6"],
        "SYNTHETIC_MARK": [SYNTHETIC_LABEL, SYNTHETIC_LABEL]
    })

    # Test Sales Slice
    df_sales = engine.slice_horizontal_blocks(df_raw, "MEESHO_GST_TCS_SALES_CURRENT")
    assert "sub_order_num" in df_sales.columns
    assert "sub_order_num.1" not in df_sales.columns
    assert df_sales.loc[0, "sub_order_num"] == "123"

    # Test Returns Slice
    df_returns = engine.slice_horizontal_blocks(df_raw, "MEESHO_GST_TCS_SALES_RETURN_CURRENT")
    assert "sub_order_num" in df_returns.columns # Suffix stripped!
    assert "sub_order_num.1" not in df_returns.columns
    assert df_returns.loc[0, "sub_order_num"] == "CN-123"
    assert df_returns.loc[0, "total_taxable_sale_value"] == "-10.0"


def test_payout_combined_headers_and_row3_skip():
    """Verify that Row 1 group + Row 2 columns are combined, and Row 3 formulas are skipped."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")
    contract = MEESHO_CONTRACTS[2] # Payout contract

    df_raw = pd.DataFrame([
        # Row 1 (merged groups)
        ["Order Details", "Order Details", "Payment Details", "Payment Details", "Deductions", "Deductions"],
        # Row 2 (columns)
        ["Sub Order No", "Supplier SKU", "Payment Date", "Final Settlement Amount", "Meesho Commission (Incl. GST)", "TCS"],
        # Row 3 (formulas to skip)
        ["Formula=1", "Formula=2", "Formula=3", "Formula=4", "Formula=5", "Formula=6"],
        # Row 4 (data 1)
        ["sub_999", "sku_test", "2026-06-01", "495.00", "5.00", "0.50"],
        # Row 5 (data 2)
        ["sub_888", "sku_test2", "2026-06-01", "190.00", "10.00", "1.00"],
    ])

    df_parsed = engine.parse_payment_order_payments(df_raw, contract)
    
    # Assert headers are concatenated
    expected_cols = [
        "Order Details_Sub Order No",
        "Order Details_Supplier SKU",
        "Payment Details_Payment Date",
        "Payment Details_Final Settlement Amount",
        "Deductions_Meesho Commission (Incl. GST)",
        "Deductions_TCS"
    ]
    assert list(df_parsed.columns) == expected_cols

    # Assert row 3 (Formula row) is skipped
    assert len(df_parsed) == 2
    assert df_parsed.loc[0, "Order Details_Sub Order No"] == "sub_999"
    assert df_parsed.loc[1, "Order Details_Sub Order No"] == "sub_888"


def test_empty_sheets_and_warning_flow():
    """Verify that empty sheets with standard warning labels do not crash and mark status warning."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")
    
    # Mock database insert and update calls
    db_mock.execute = MagicMock()
    db_mock.commit = MagicMock()

    # Create payout report but with empty optional sheets and warning labels
    sheets_data = {
        "Order Payments": [
            ["Order Details", "Order Details", "Payment Details", "Payment Details"],
            ["Sub Order No", "Supplier SKU", "Payment Date", "Final Settlement Amount"],
            ["Formula", "Formula", "Formula", "Formula"],
            ["sub_100", "sku_100", "2026-06-01", "450.00"]
        ],
        "Ads Cost": [
            ["No data is available for these dates", "", ""], # Standard empty sheet label
            ["Campaign ID", "Campaign Name", "Amount"]
        ],
        "Referral Payments": [] # Empty sheet
    }
    
    content = create_synthetic_excel(sheets_data)
    
    # Mock uniqueness check
    db_mock.execute.return_value.fetchone.return_value = None # No existing file
    
    res = engine.process_and_save(content, "SP_ORDER_ADS_REFERRAL_PAYMENT_FILE_PREVIOUS_PAYMENT.xlsx", "2026-06")
    
    assert res["validation_status"] == "WARNING"
    assert "Ads Cost" in res["empty_sheets"]
    assert "Referral Payments" in res["empty_sheets"]
    assert res["records_inserted"] == 1


def test_missing_required_columns():
    """Verify that validation fails if required contract columns are missing."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")
    contract = MEESHO_CONTRACTS[2]

    # Data missing "Payment Details_Final Settlement Amount"
    df_missing = pd.DataFrame(columns=[
        "Order Details_Sub Order No",
        "Payment Details_Payment Date",
        "Order Details_Supplier SKU"
    ])

    errors = engine.run_pre_ingestion_validation(df_missing, contract, "Order Payments")
    assert len(errors) > 0
    assert any("Final Settlement Amount" in err for err in errors)


def test_amount_sign_rules():
    """Verify that claim event signs match the contract amount_sign_rules_json config."""
    db_mock = MagicMock(spec=Session)
    engine = MeeshoExtractionEngine(db_mock, "user-123")
    contract = MEESHO_CONTRACTS[2] # Payout contract

    # Sign rules contain: recovery: -1, claims: 1, gst_compensation: 1, waivers: 1, compensation: 1
    # 1. Recovery
    df_rec = pd.DataFrame({
        "Sub Order No": ["sub_r"],
        "Payment Date": ["2026-06-01"],
        "Type": ["Logistics Recovery fee"],
        "Reason": ["Weight overcharge penalty"],
        "Amount": ["150.00"]
    })
    
    db_mock.execute = MagicMock()
    engine.save_compensation_recovery(df_rec, contract, MagicMock())
    
    # Retrieve the values passed to insert
    args = db_mock.execute.call_args[0][1]
    assert args["claim_type"] == "RECOVERY"
    assert args["amt"] == -150.0 # Multiplied by -1!

    # 2. Compensation
    df_comp = pd.DataFrame({
        "Sub Order No": ["sub_c"],
        "Payment Date": ["2026-06-01"],
        "Type": ["Lost Cargo Compensation"],
        "Reason": ["Lost in transit"],
        "Amount": ["500.00"]
    })
    
    db_mock.execute = MagicMock()
    engine.save_compensation_recovery(df_comp, contract, MagicMock())
    
    args = db_mock.execute.call_args[0][1]
    assert args["claim_type"] == "COMPENSATION"
    assert args["amt"] == 500.0 # Multiplied by 1!
